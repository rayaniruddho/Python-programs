# -*- coding: utf-8 -*-
"""
Created on Wed Jul  5 09:46:26 2017

@author: Aniruddho
"""

import openpyxl
import itertools
from collections import Counter

wb = openpyxl.load_workbook('data.xlsx')
wb1 = openpyxl.load_workbook('output.xlsx')

sheet = wb.get_sheet_by_name('Input')
sheet1 = wb1.get_sheet_by_name('Sheet1')

for i in range(1,7):
    _cell = sheet1.cell(row = 1,column = i)
    _cell.font = _cell.font.copy(bold=True)

s = 0
age_values,income_values,student_values,credit_values,label_values=[],[],[],[],[]
a1,a2,a3,a4 = [],[],[],[]
in_data=[]

for row in range(2, sheet.max_row + 1):
       # Each row in the spreadsheet has data for one type of customer.
       s+=1
       
       age  = sheet['B' + str(row)].value
       age_values.append(age)         
           
       income = sheet['C' + str(row)].value
       income_values.append(income)
                 
       student = sheet['D' + str(row)].value
       student_values.append(student)
                     
       credit = sheet['E' + str(row)].value
       credit_values.append(credit)
                    
       label = sheet['F' + str(row)].value
       label_values.append(label)
       
       in_row = (age,income,student,credit)
       in_data.append(in_row)
       
       s_age1 = (label,age)
       a1.append(s_age1)
       
       s_income1 = (label,income)
       a2.append(s_income1)
       
       s_student1 = (label,student)
       a3.append(s_student1)
       
       s_credit1 = (label,credit)
       a4.append(s_credit1)
           
s_cl = Counter(label_values)

s_age = Counter(age_values)
s_cl_age = Counter(a1)

s_income = Counter(income_values)
s_cl_income = Counter(a2)

s_student = Counter(student_values)
s_cl_student = Counter(a3)

s_credit = Counter(credit_values)
s_cl_credit = Counter(a4)

s_complete = s_cl_age + s_cl_income + s_cl_student + s_cl_credit

rows = len(s_age) * len(s_income) * len(s_student) * len(s_credit)
sheet1.cell(row=1,column=1).value = 'RID'
sheet1.cell(row=1,column=2).value = 'AGE'
sheet1.cell(row=1,column=3).value = 'INCOME'
sheet1.cell(row=1,column=4).value = 'STUDENT'
sheet1.cell(row=1,column=5).value = 'CREDIT RATING'
sheet1.cell(row=1,column=6).value = 'CLASS LABEL=BUYS COMPUTER'

data_set = list(itertools.product(list(s_age.keys()),list(s_income.keys()),
                                  list(s_student.keys()),list(s_credit.keys())))

data_set = list(set(data_set) - set(in_data))

for i in range(2,len(data_set)+2):
    for j in range(1,6):
        if j == 1:
            sheet1.cell(row=i,column=1).value = i-1
        elif j == 2:
            sheet1.cell(row=i,column=2).value = data_set[i-2][0]
        elif j == 3:
            sheet1.cell(row=i,column=3).value = data_set[i-2][1]
        elif j == 4:
            sheet1.cell(row=i,column=4).value = data_set[i-2][2]
        elif j == 5:
            sheet1.cell(row=i,column=5).value = data_set[i-2][3]

for i in range(2,sheet1.max_row + 1): 
    curr = 0
    for k in s_cl:
        p_x_cl = 1        
        for j in range(2,6):
            x = (k,sheet1.cell(row=i,column=j).value)
            p_x_cl *= s_complete[x]/s_cl[k]
        p_x_cl *= (s_cl[k]/s)
        
        if p_x_cl > curr:
            curr = p_x_cl
            cl = k
            #print(curr,k)
    sheet1.cell(row=i,column=6).value = cl
wb1.save('output.xlsx')
